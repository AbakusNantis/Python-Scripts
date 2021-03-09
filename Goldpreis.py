import requests
from bs4 import BeautifulSoup
from datetime import date, datetime
from openpyxl import Workbook, load_workbook
#import openpyxl

#Münzen
url_mz = 'https://shop.degussa-goldhandel.de/gold/anlagemuenzen/maple-leaf'

#Maple 0.5 oz, 1 oz
mz = {
      'product-price-89',
      'product-price-88',      
      }

#Barren 40g, 50g, 100g
url_br = 'https://shop.degussa-goldhandel.de/gold/goldbarren/investmentbarren'
br ={
     'product-price-2413',
     'product-price-2187',
     'product-price-1841',
     }


#Create List
liste =[]

#Append date and time
today_date = date.today().strftime('%d.%m.%Y')
liste.append(today_date)
today_time = datetime.now().strftime('%H:%M:%S')
liste.append(today_time)

def adder(url, product_list):
    r = requests.get(url)
    print(r.status_code)
    for m in product_list:
        
        soup = BeautifulSoup(r.text, 'lxml')
        fen = soup.find('span', class_='regular-price', id=m)
        prc = fen.find('span', class_='price').text
        prc = float(prc.replace('\xa0€','').replace('.','').replace(',','.'))
        liste.append(prc)

adder(url_mz, mz)
adder(url_br, br)

print(liste)

#Adding to Excel
url_xl = 'C:\\Users\\basti\\OneDrive\Ausarbeitung\\03 - Finanzen etc\\09 - Aktien\\'
file_xl = 'GoldpreisBeobachtung.xlsx'

wb = load_workbook(url_xl + file_xl)
ws = wb['Degussa']

for cell in ws.iter_cols(max_col = 1, values_only = True):
    nb_row = len(cell)

from string import ascii_uppercase

char_row = 'A'
alp = ascii_uppercase

for i in range(len(liste)):
    # Wie kann man den Buchstaben durch Iterieren bestimmen?
    alph = ascii_uppercase
    col_lt = alph[i]
    #+1, weil so die Zahl die letzte Zeile gewählt wird
    cell_nr = col_lt +str(nb_row+1)
    cell = ws[cell_nr]
    cell.value = liste[i]

wb.save(url_xl + file_xl)
